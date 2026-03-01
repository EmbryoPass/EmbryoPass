import io
import os
import smtplib
import secrets
import string
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from sqlalchemy import text

GMAIL_USER = os.environ.get('GMAIL_USER', 'museoembriologia@gmail.com')
GMAIL_PASSWORD = os.environ.get('GMAIL_PASSWORD')

NOMBRE_MUSEO = '"Museo de Embriología Dra. Dora Virginia Chávez Corral"'


def enviar_correo(destinatario, asunto, cuerpo_html):
    """Envía un correo HTML usando Gmail SMTP."""
    if not GMAIL_USER or not GMAIL_PASSWORD:
        print("⚠️ GMAIL_USER/GMAIL_PASSWORD no configurados; no se envía correo.")
        return

    mensaje = MIMEMultipart()
    mensaje['From'] = GMAIL_USER
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto
    mensaje.attach(MIMEText(cuerpo_html, 'html'))

    servidor = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
    try:
        servidor.starttls()
        servidor.login(GMAIL_USER, GMAIL_PASSWORD)
        servidor.sendmail(GMAIL_USER, destinatario, mensaje.as_string())
    finally:
        servidor.quit()


def enviar_correo_con_excel(destinatario, asunto, cuerpo_html, nombre_archivo_excel,
                            datos_grupo=None):
    """Envía un correo HTML con un Excel adjunto para llenar datos de estudiantes.

    datos_grupo (dict opcional) con claves:
        institucion, nivel, ciudad, estado, fecha, encargado
    Si se pasa, se agrega una sección de encabezado en el Excel con esos datos
    para que el encargado no tenga que repetirlos por alumno.
    """
    if not GMAIL_USER or not GMAIL_PASSWORD:
        print("⚠️ GMAIL_USER/GMAIL_PASSWORD no configurados; no se envía correo.")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Asistencia"

    cafe        = PatternFill("solid", fgColor="8B4513")
    cafe_claro  = PatternFill("solid", fgColor="D4956A")
    gris        = PatternFill("solid", fgColor="F2EBE3")
    bold_white  = Font(bold=True, color="FFFFFF", size=11)
    bold_dark   = Font(bold=True, color="3B2008", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")

    fila = 1  # cursor de fila

    # ── Sección de datos del grupo (pre-llenada) ──────────────────────────────
    if datos_grupo:
        # Título de sección
        ws.merge_cells(f"A{fila}:D{fila}")
        c = ws.cell(fila, 1, "DATOS DEL GRUPO")
        c.font = bold_white; c.fill = cafe; c.alignment = center
        ws.row_dimensions[fila].height = 20
        fila += 1

        campos = [
            ("Institución",     datos_grupo.get("institucion", "")),
            ("Nivel académico", datos_grupo.get("nivel", "")),
            ("Ciudad",          datos_grupo.get("ciudad", "")),
            ("Estado",          datos_grupo.get("estado", "")),
            ("Fecha de visita", datos_grupo.get("fecha", "")),
            ("Encargado",       datos_grupo.get("encargado", "")),
        ]
        for etiqueta, valor in campos:
            ws.merge_cells(f"A{fila}:D{fila}")
            c_lbl = ws.cell(fila, 1)
            c_lbl.fill = gris
            c_lbl.alignment = left
            c_lbl.font = Font(bold=False, color="3B2008", size=10)
            c_lbl.value = f"  {etiqueta}:   {valor}"
            ws.row_dimensions[fila].height = 18
            fila += 1

        # Fila en blanco separadora
        fila += 1

    # ── Encabezados de la tabla de alumnos ────────────────────────────────────
    # NOTA IMPORTANTE PARA EL PARSER: la fila de encabezados siempre
    # tiene exactamente estos 4 valores; se usa para detectar dónde
    # empiezan los datos al procesar el Excel subido.
    encabezados_tabla = ["No.", "Nombre completo", "Edad", "Sexo (Hombre/Mujer)"]
    for col_idx, titulo in enumerate(encabezados_tabla, start=1):
        c = ws.cell(fila, col_idx, titulo)
        c.font = bold_white; c.fill = cafe_claro; c.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = 28
    ws.row_dimensions[fila].height = 22
    fila += 1

    # ── Filas vacías numeradas ────────────────────────────────────────────────
    alumnos_estimados = datos_grupo.get("numero_alumnos", 20) if datos_grupo else 20
    filas_vacias = max(int(alumnos_estimados) + 5, 15)  # margen extra
    for i in range(1, filas_vacias + 1):
        ws.cell(fila, 1, i).alignment = center
        fila += 1

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Armar correo
    mensaje = MIMEMultipart()
    mensaje['From'] = GMAIL_USER
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto
    mensaje.attach(MIMEText(cuerpo_html, 'html'))

    # Adjuntar Excel
    adjunto = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    adjunto.set_payload(buffer.read())
    encoders.encode_base64(adjunto)
    adjunto.add_header('Content-Disposition', f'attachment; filename="{nombre_archivo_excel}"')
    mensaje.attach(adjunto)

    servidor = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
    try:
        servidor.starttls()
        servidor.login(GMAIL_USER, GMAIL_PASSWORD)
        servidor.sendmail(GMAIL_USER, destinatario, mensaje.as_string())
    finally:
        servidor.quit()


def generar_password_segura(longitud=14):
    """Genera una contraseña segura aleatoria."""
    alfabeto = string.ascii_letters + string.digits + "!@#$%_-"
    while True:
        pwd = ''.join(secrets.choice(alfabeto) for _ in range(longitud))
        if (any(c.islower() for c in pwd) and
                any(c.isupper() for c in pwd) and
                any(c.isdigit() for c in pwd) and
                any(c in "!@#$%_-" for c in pwd)):
            return pwd


def verificar_y_agregar_columnas(app, db):
    """Agrega columnas faltantes si no existen."""
    with app.app_context():
        inspector = db.inspect(db.engine)

        # Tabla cita
        columnas_cita = [col['name'] for col in inspector.get_columns('cita')]
        cambios_cita = []
        if 'institucion' not in columnas_cita:
            cambios_cita.append("ADD COLUMN institucion VARCHAR(100)")
        if 'nivel_educativo' not in columnas_cita:
            cambios_cita.append("ADD COLUMN nivel_educativo VARCHAR(50)")
        if 'ciudad' not in columnas_cita:
            cambios_cita.append("ADD COLUMN ciudad VARCHAR(100)")
        if 'estado_republica' not in columnas_cita:
            cambios_cita.append("ADD COLUMN estado_republica VARCHAR(100)")
        if cambios_cita:
            db.session.execute(text(f"ALTER TABLE cita {', '.join(cambios_cita)};"))
            db.session.commit()
            print("✅ Columnas agregadas a cita:", cambios_cita)

        # Tabla visita_grupal
        columnas_vg = [col['name'] for col in inspector.get_columns('visita_grupal')]
        cambios_vg = []
        if 'ciudad' not in columnas_vg:
            cambios_vg.append("ADD COLUMN ciudad VARCHAR(100)")
        if 'estado_republica' not in columnas_vg:
            cambios_vg.append("ADD COLUMN estado_republica VARCHAR(100)")
        if 'bachillerato' not in columnas_vg:
            cambios_vg.append("ADD COLUMN bachillerato VARCHAR(150)")
        if cambios_vg:
            db.session.execute(text(f"ALTER TABLE visita_grupal {', '.join(cambios_vg)};"))
            db.session.commit()
            print("✅ Columnas agregadas a visita_grupal:", cambios_vg)
