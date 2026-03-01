import pytz
import pandas as pd
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, make_response

from app import db
from app.models import Cita, Horario, VisitaGrupal, AdminSecret
from app.utils import enviar_correo, enviar_correo_con_excel, generar_password_segura, GMAIL_USER, NOMBRE_MUSEO

admin_bp = Blueprint('admin', __name__)

ENCARGADO_USER = 'admin'
ENCARGADO_PASS = '1234'
URL_SITIO = 'https://quixotic-veronika-uach-98c1e80d.koyeb.app'


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario' not in session:
            flash('⚠️ Debes iniciar sesión primero.', 'warning')
            return redirect(url_for('admin.login'))
        return f(*args, **kwargs)
    return decorated


@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']
        if usuario == ENCARGADO_USER and password == ENCARGADO_PASS:
            session['usuario'] = usuario
            return redirect(url_for('admin.dashboard'))
        flash('❌ Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')


@admin_bp.route('/logout')
def logout():
    session.pop('usuario', None)
    flash('✅ Sesión cerrada.', 'success')
    return redirect(url_for('admin.login'))


@admin_bp.route('/dashboard')
@login_required
def dashboard():
    zona = pytz.timezone('America/Chihuahua')
    ahora = datetime.now(zona)
    rango = request.args.get('rango', default='30')
    tipo = request.args.get('tipo', default='todas')

    if rango == '7':
        inicio_rango = ahora - timedelta(days=7)
    elif rango == '30':
        inicio_rango = ahora - timedelta(days=30)
    elif rango == 'mes':
        inicio_rango = ahora.replace(day=1)
    elif rango == 'todo':
        inicio_rango = datetime.min.replace(tzinfo=zona)
    else:
        inicio_rango = ahora - timedelta(days=30)

    historial_completo = []

    for c in Cita.query.all():
        try:
            fecha = datetime.strptime(c.fecha_hora, "%d/%m/%Y %I:%M %p")
        except ValueError:
            try:
                fecha = datetime.strptime(c.fecha_hora, "%Y-%m-%d %H:%M")
            except ValueError:
                continue
        fecha = zona.localize(fecha)

        if fecha < ahora and fecha >= inicio_rango and c.estado != 'cancelada':
            historial_completo.append({
                'tipo': 'Individual', 'id': c.id, 'nombre': c.nombre,
                'edad': c.edad, 'sexo': c.sexo, 'correo': c.correo,
                'telefono': c.telefono, 'fecha_hora': c.fecha_hora,
                'estado': c.estado, 'asistio': c.asistio if c.asistio in ['sí', 'no'] else None,
                'institucion': c.institucion, 'nivel': c.nivel_educativo,
                'ciudad': c.ciudad, 'estado_rep': c.estado_republica,
            })

    # ── Horarios futuros con sus citas activas embebidas ─────────────────────
    horarios = []
    for h in Horario.query.all():
        try:
            fecha = datetime.strptime(h.fecha_hora, "%d/%m/%Y %I:%M %p")
        except ValueError:
            fecha = datetime.strptime(h.fecha_hora, "%Y-%m-%d %H:%M")
        fecha = zona.localize(fecha)

        if fecha >= ahora:
            citas_activas = Cita.query.filter_by(fecha_hora=h.fecha_hora, estado='activa').all()
            total = h.disponibles + len(citas_activas)
            horarios.append({
                'id':          h.id,
                'fecha_hora':  fecha.strftime("%d/%m/%Y %I:%M %p"),
                'disponibles': h.disponibles,
                'total':       total,
                'citas':       citas_activas,
            })

    horarios.sort(key=lambda x: datetime.strptime(x['fecha_hora'], "%d/%m/%Y %I:%M %p"))

    # ── Historial de visitas grupales (fecha confirmada en el pasado) ─────────
    historial_grupales = []
    for v in VisitaGrupal.query.filter(
        VisitaGrupal.fecha_confirmada.isnot(None),
        VisitaGrupal.estado == 'aceptada'
    ).order_by(VisitaGrupal.id.desc()).all():
        try:
            fecha_v = datetime.strptime(v.fecha_confirmada, "%d/%m/%Y %I:%M %p")
        except ValueError:
            try:
                fecha_v = datetime.strptime(v.fecha_confirmada, "%d/%m/%Y %H:%M")
            except ValueError:
                continue
        fecha_v = zona.localize(fecha_v)
        if fecha_v < ahora and fecha_v >= inicio_rango:
            historial_grupales.append(v)

    visitas_grupales = VisitaGrupal.query.order_by(VisitaGrupal.id.desc()).all()

    secret = AdminSecret.query.get(1)
    admin_password = secret.password if secret else None
    admin_password_at = None
    if secret and secret.created_at:
        chih = pytz.timezone('America/Chihuahua')
        admin_password_at = secret.created_at.replace(tzinfo=pytz.utc).astimezone(chih).strftime("%d/%m/%Y %I:%M %p")

    return render_template(
        'dashboard.html',
        historial_completo=historial_completo,
        historial_grupales=historial_grupales,
        horarios=horarios,
        rango=rango,
        admin_password=admin_password,
        admin_password_at=admin_password_at,
        tipo=tipo,
        tipo_filtro=tipo,
        visitas_grupales=visitas_grupales
    )


@admin_bp.route('/marcar_asistencia/<int:id_cita>/<estado>')
@login_required
def marcar_asistencia(id_cita, estado):
    cita = Cita.query.get(id_cita)
    if cita and estado in ['sí', 'no']:
        cita.asistio = estado
        db.session.commit()
        flash('✅ Asistencia registrada.', 'success')
    else:
        flash('❌ Error al actualizar asistencia.', 'danger')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/cancelar_cita/<int:id_cita>')
@login_required
def cancelar_cita(id_cita):
    cita = Cita.query.get(id_cita)
    if not cita:
        flash('❌ Cita no encontrada.', 'danger')
        return redirect(url_for('admin.dashboard'))

    horario = Horario.query.filter_by(fecha_hora=cita.fecha_hora).first()
    cita.estado = 'cancelada'
    if horario:
        horario.disponibles += 1
    db.session.commit()

    try:
        enviar_correo(cita.correo, f'Cancelación de Cita – {NOMBRE_MUSEO}', f"""
<html><body style="font-family:Arial,sans-serif;color:#333;">
  <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #f5c6cb;border-radius:10px;">
    <h2 style="color:#d9534f;">Cancelación de Cita</h2>
    <p>Hola <strong>{cita.nombre}</strong>,</p>
    <p>Tu cita al {NOMBRE_MUSEO} del <strong>{cita.fecha_hora}</strong> ha sido cancelada
       debido a un imprevisto.</p>
    <p>Puedes agendar una nueva cita cuando lo desees.</p>
    <div style="text-align:center;margin-top:20px;">
      <a href="{URL_SITIO}/agendar-cita"
         style="background:#4a90e2;color:white;padding:12px 24px;
                text-decoration:none;border-radius:6px;font-weight:bold;
                display:inline-block;">
        Agendar nueva cita
      </a>
    </div>
    <p style="margin-top:16px;">Gracias por tu comprensión.</p>
  </div>
</body></html>""")
        flash('✅ Cita cancelada, correo enviado y espacio liberado.', 'success')
    except Exception as e:
        print(f"[EMAIL] Error: {e}")
        flash('✅ Cita cancelada y espacio liberado (correo no enviado).', 'warning')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/eliminar_cita/<int:id_cita>')
@login_required
def eliminar_cita(id_cita):
    cita = Cita.query.get(id_cita)
    if cita:
        db.session.delete(cita)
        db.session.commit()
        flash('✅ Cita eliminada.', 'success')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/agregar_horario', methods=['POST'])
@login_required
def agregar_horario():
    fecha_hora = request.form['fecha_hora']
    disponibles = int(request.form['disponibles'])
    if disponibles > 10:
        disponibles = 10
    db.session.add(Horario(fecha_hora=fecha_hora, disponibles=disponibles))
    db.session.commit()
    flash('✅ Horario agregado.', 'success')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/eliminar_horario/<int:id_horario>')
@login_required
def eliminar_horario(id_horario):
    # 1. Localizar el horario por su ID único (esto nunca falla)
    horario = Horario.query.get(id_horario)
    if not horario:
        flash('❌ Horario no encontrado.', 'danger')
        return redirect(url_for('admin.dashboard'))

    # 2. Limpieza de seguridad: quitamos espacios extras de la fecha para la búsqueda
    fecha_busqueda = horario.fecha_hora.strip()

    # 3. Buscar citas que coincidan con esa fecha exacta
    # Usamos .strip() en la base de datos para asegurar que coincidan sin errores de espacios
    citas = Cita.query.filter(Cita.fecha_hora.contains(fecha_busqueda)).all()
    
    for c in citas:
        if c.estado == 'activa':
            c.estado = 'cancelada'

    try:
        # 4. Borrar el horario y confirmar cambios
        db.session.delete(horario)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR CRÍTICO]: {str(e)}")
        flash('❌ Error al eliminar: Problema de sincronización en la base de datos.', 'danger')
        return redirect(url_for('admin.dashboard'))

    # 5. Notificaciones (Tus plantillas originales sin cambios)
    errores = 0
    for c in citas:
        if c.estado == 'cancelada':
            try:
                enviar_correo(c.correo, f'Cancelación de Cita – {NOMBRE_MUSEO}', f"""
<html><body style="font-family:Arial,sans-serif;color:#333;">
  <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #f5c6cb;border-radius:10px;">
    <h2 style="color:#d9534f;">Cancelación de Cita</h2>
    <p>Hola <strong>{c.nombre}</strong>,</p>
    <p>Tu cita del <strong>{c.fecha_hora}</strong> ha sido cancelada por cambios de
       disponibilidad en el {NOMBRE_MUSEO}.</p>
    <p>Puedes agendar una nueva cita cuando lo desees.</p>
    <div style="text-align:center;margin-top:20px;">
      <a href="{URL_SITIO}/agendar-cita"
         style="background:#4a90e2;color:white;padding:12px 24px;
                text-decoration:none;border-radius:6px;font-weight:bold;
                display:inline-block;">
        Agendar nueva cita
      </a>
    </div>
    <p style="margin-top:16px;">Gracias por tu comprensión.</p>
  </div>
</body></html>""")
            except Exception as env_err:
                print(f"[EMAIL ERROR]: {env_err}")
                errores += 1

    if errores:
        flash(f'✅ Horario eliminado. {len(citas)} cita(s) cancelada(s) ({errores} fallos de envío).', 'warning')
    else:
        flash('✅ Horario eliminado y base de datos actualizada.', 'success')
        
    return redirect(url_for('admin.dashboard'))
    

def _nivel_str(visita):
    if visita.nivel == 'Bachillerato' and visita.bachillerato:
        return f"{visita.nivel} – {visita.bachillerato}"
    return visita.nivel or '—'


@admin_bp.route('/aceptar_visita/<int:id>')
@login_required
def aceptar_visita(id):
    visita = VisitaGrupal.query.get(id)
    if not visita:
        flash('❌ Visita no encontrada.', 'danger')
        return redirect(url_for('admin.dashboard'))

    ya_aceptada = (visita.estado == 'aceptada')
    visita.estado = 'aceptada'
    db.session.commit()

    if not ya_aceptada and visita.correo:
        try:
            enviar_correo(visita.correo, f'Solicitud de visita grupal aceptada — {NOMBRE_MUSEO}', f"""
<html><body style="font-family:Arial,sans-serif;color:#333;">
  <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #eee;border-radius:10px;">
    <h2 style="color:#4a90e2;">Solicitud aceptada</h2>
    <p>Hola <strong>{visita.encargado}</strong>,</p>
    <p>Tu solicitud de visita grupal al {NOMBRE_MUSEO} ha sido <strong>aceptada</strong>. En breve recibirás un correo con la fecha y hora confirmadas.</p>
    <ul style="line-height:1.6;">
      <li><strong>Institución / Plantel:</strong> {visita.institucion}</li>
      <li><strong>Nivel académico:</strong> {_nivel_str(visita)}</li>
      <li><strong>Alumnos estimados:</strong> {visita.numero_alumnos}</li>
    </ul>
    <p>Gracias por tu interés en el {NOMBRE_MUSEO}.</p>
  </div>
</body></html>""")
            flash('✅ Visita aceptada y correo enviado.', 'success')
        except Exception as e:
            print(f"[EMAIL] Error: {e}")
            flash('✅ Visita aceptada, pero no se pudo enviar el correo.', 'warning')
    else:
        flash('ℹ️ La visita ya estaba aceptada.', 'info')

    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/rechazar_visita/<int:id>')
@login_required
def rechazar_visita(id):
    visita = VisitaGrupal.query.get(id)
    if not visita:
        flash('❌ Visita no encontrada.', 'danger')
        return redirect(url_for('admin.dashboard'))

    visita.estado = 'rechazada'
    db.session.commit()

    try:
        enviar_correo(visita.correo, f'Solicitud de visita grupal rechazada — {NOMBRE_MUSEO}', f"""
<html><body style="font-family:Arial,sans-serif;color:#333;">
  <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #eee;border-radius:10px;">
    <h2 style="color:#d9534f;">Solicitud rechazada</h2>
    <p>Hola <strong>{visita.encargado}</strong>,</p>
    <p>Lamentamos informarte que tu solicitud de visita grupal al {NOMBRE_MUSEO} ha sido <strong>rechazada</strong>.</p>
    <p>Si lo deseas, puedes presentar una nueva solicitud con otras fechas.</p>
    <div style="text-align:center;margin-top:20px;">
      <a href="{URL_SITIO}/solicitar-visita-grupal"
         style="background:#4a90e2;color:white;padding:12px 24px;
                text-decoration:none;border-radius:6px;font-weight:bold;
                display:inline-block;">
        Solicitar nueva visita grupal
      </a>
    </div>
    <p style="margin-top:16px;">Gracias por tu interés en el {NOMBRE_MUSEO}.</p>
  </div>
</body></html>""")
        flash('✅ Visita rechazada y correo enviado.', 'success')
    except Exception as e:
        print(f"[EMAIL] Error: {e}")
        flash('⚠️ Visita rechazada, pero no se pudo enviar el correo.', 'warning')

    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/asignar_fecha_visita/<int:id>', methods=['POST'])
@login_required
def asignar_fecha_visita(id):
    visita = VisitaGrupal.query.get(id)
    if not visita:
        flash('❌ Visita no encontrada.', 'danger')
        return redirect(url_for('admin.dashboard'))

    if visita.estado != 'aceptada':
        flash('❌ La visita aún no ha sido aceptada.', 'danger')
        return redirect(url_for('admin.dashboard'))

    fecha = (request.form.get('fecha_confirmada') or '').strip()
    if not fecha:
        flash('❌ Debes proporcionar una fecha válida.', 'danger')
        return redirect(url_for('admin.dashboard'))

    fecha_anterior = (visita.fecha_confirmada or '').strip()
    visita.fecha_confirmada = fecha
    db.session.commit()

    try:
        nombre_excel = f"Lista_estudiantes_{visita.institucion.replace(' ','_')}_{fecha.replace('/','-').replace(' ','_')}.xlsx"
        
        # Preparamos los datos del grupo para que aparezcan en el Excel en ambos casos
        datos_grupo = {
            'institucion':    visita.institucion,
            'nivel':          _nivel_str(visita),
            'ciudad':         visita.ciudad or '—',
            'estado':         visita.estado_republica or '—',
            'fecha':          fecha,
            'encargado':      visita.encargado,
            'numero_alumnos': visita.numero_alumnos,
        }

        if not fecha_anterior:
            # PRIMERA CONFIRMACIÓN
            cuerpo = f"""
<html><body style="font-family:Arial,sans-serif;color:#333;">
  <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #eee;border-radius:10px;">
    <h2 style="color:#4a90e2;">Confirmación de visita grupal – {NOMBRE_MUSEO}</h2>
    <p>Hola <strong>{visita.encargado}</strong>,</p>
    <p>Tu solicitud de visita grupal al {NOMBRE_MUSEO} ha sido <strong>confirmada</strong> para:</p>
    <p style="font-size:18px;font-weight:bold;margin:12px 0;">{fecha}</p>
    <ul style="line-height:1.6;">
      <li><strong>Institución / Plantel:</strong> {visita.institucion}</li>
      <li><strong>Nivel académico:</strong> {_nivel_str(visita)}</li>
      <li><strong>Alumnos estimados:</strong> {visita.numero_alumnos}</li>
    </ul>
    <p><strong>Duración estimada de la visita:</strong> 10 a 15 minutos.</p>
    <p><strong>Indicaciones durante la visita:</strong></p>
    <ul style="line-height:1.6;">
      <li>No tocar las exhibiciones.</li>
      <li>No comer ni beber dentro del museo.</li>
      <li>No hablar en voz alta.</li>
      <li>No tomar fotos ni videos.</li>
      <li>No correr ni empujar dentro del museo.</li>
      <li>No manipular etiquetas, carteles o información sobre las piezas.</li>
    </ul>
    <p>📎 Se adjunta una lista en Excel para que registres los datos de los estudiantes que asistirán.</p>
    <p style="background:#fff3cd;border:1px solid #ffc107;border-radius:6px;padding:12px;margin:16px 0;">
      <strong style="font-size:14px;">
        ⚠️ IMPORTANTE: AL DEVOLVER ESTE ARCHIVO, ENVÍELO ÚNICAMENTE EN FORMATO
        EXCEL (.xlsx). NO LO CONVIERTA A PDF NI A NINGÚN OTRO FORMATO,
        YA QUE DE LO CONTRARIO NO PODRÁ SER PROCESADO.
      </strong>
    </p>
    <p>Le recomendamos llegar 15 minutos antes del horario programado.</p>
    <p>Gracias por tu interés en el {NOMBRE_MUSEO}.</p>
  </div>
</body></html>"""
            # Se agregó 'datos_grupo=datos_grupo' aquí también
            enviar_correo_con_excel(visita.correo, f'Confirmación de visita grupal — {NOMBRE_MUSEO}', cuerpo, nombre_excel, datos_grupo=datos_grupo)
            flash('📅 Fecha confirmada y correo con Excel adjunto enviado.', 'success')

        elif fecha_anterior != fecha:
            # ACTUALIZACIÓN / REPROGRAMACIÓN
            cuerpo_reagenda = f"""
<html><body style="font-family:Arial,sans-serif;color:#333;">
  <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #eee;border-radius:10px;">
    <h2 style="color:#4a90e2;">Actualización de fecha – visita grupal – {NOMBRE_MUSEO}</h2>
    <p>Hola <strong>{visita.encargado}</strong>,</p>
    <p>La fecha de tu visita grupal ha sido <strong>actualizada</strong>:</p>
    <ul style="line-height:1.6;">
      <li><strong>Fecha anterior:</strong> {fecha_anterior}</li>
      <li><strong>Nueva fecha confirmada:</strong> <strong style="font-size:16px;">{fecha}</strong></li>
      <li><strong>Institución / Plantel:</strong> {visita.institucion}</li>
      <li><strong>Nivel académico:</strong> {_nivel_str(visita)}</li>
      <li><strong>Alumnos estimados:</strong> {visita.numero_alumnos}</li>
    </ul>
    <p><strong>Duración estimada de la visita:</strong> 10 a 15 minutos.</p>
    <p><strong>Indicaciones durante la visita:</strong></p>
    <ul style="line-height:1.6;">
      <li>No tocar las exhibiciones.</li>
      <li>No comer ni beber dentro del museo.</li>
      <li>No hablar en voz alta.</li>
      <li>No tomar fotos ni videos.</li>
      <li>No correr ni empujar dentro del museo.</li>
      <li>No manipular etiquetas, carteles o información sobre las piezas.</li>
    </ul>
    <p>📎 Se adjunta nuevamente la lista en Excel actualizada con la nueva fecha.
       Por favor, llénala con los datos de los estudiantes que asistirán.</p>
    <p style="background:#fff3cd;border:1px solid #ffc107;border-radius:6px;padding:12px;margin:16px 0;">
      <strong style="font-size:14px;">
        ⚠️ IMPORTANTE: AL DEVOLVER ESTE ARCHIVO, ENVÍELO ÚNICAMENTE EN FORMATO
        EXCEL (.xlsx). NO LO CONVIERTA A PDF NI A NINGÚN OTRO FORMATO,
        YA QUE DE LO CONTRARIO NO PODRÁ SER PROCESADO.
      </strong>
    </p>
    <p>Le recomendamos llegar 15 minutos antes del horario programado.</p>
    <p>Si hay algún inconveniente, responde a este correo para coordinar.</p>
    <p>Gracias por tu interés en el {NOMBRE_MUSEO}.</p>
  </div>
</body></html>"""
            enviar_correo_con_excel(
                visita.correo,
                f'Actualización de fecha — {NOMBRE_MUSEO}',
                cuerpo_reagenda, nombre_excel, datos_grupo=datos_grupo
            )
            flash('📅 Fecha actualizada y nuevo correo con Excel enviado.', 'success')
        else:
            flash('📅 Fecha confirmada.', 'success')

    except Exception as e:
        print(f"[EMAIL] Error: {e}")
        flash('📅 Fecha guardada, pero no se pudo enviar el correo.', 'warning')

    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/cancelar_visita_grupal/<int:id>')
@login_required
def cancelar_visita_grupal(id):
    visita = VisitaGrupal.query.get(id)
    if not visita:
        flash('❌ Visita no encontrada.', 'danger')
        return redirect(url_for('admin.dashboard'))

    if visita.estado in ['rechazada', 'cancelada']:
        flash('ℹ️ No es posible cancelar una solicitud rechazada o ya cancelada.', 'info')
        return redirect(url_for('admin.dashboard'))

    visita.estado = 'cancelada'
    db.session.commit()

    try:
        enviar_correo(visita.correo, f'Cancelación de visita grupal – {NOMBRE_MUSEO}', f"""
<html><body style="font-family:Arial,sans-serif;color:#333;">
  <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #eee;border-radius:10px;">
    <h2 style="color:#d9534f;">Cancelación de visita grupal</h2>
    <p>Hola <strong>{visita.encargado}</strong>,</p>
    <p>Tu solicitud de visita grupal al {NOMBRE_MUSEO} ha sido <strong>cancelada</strong>.</p>
    <p>Puedes solicitar una nueva visita cuando lo desees.</p>
    <div style="text-align:center;margin-top:20px;">
      <a href="{URL_SITIO}/solicitar-visita-grupal"
         style="background:#4a90e2;color:white;padding:12px 24px;
                text-decoration:none;border-radius:6px;font-weight:bold;
                display:inline-block;">
        Solicitar nueva visita grupal
      </a>
    </div>
    <p style="margin-top:16px;">Gracias por tu interés en el {NOMBRE_MUSEO}.</p>
  </div>
</body></html>""")
        flash('✅ Visita cancelada y notificación enviada.', 'success')
    except Exception as e:
        print(f"[EMAIL] Error: {e}")
        flash('✅ Visita cancelada, pero no se pudo enviar el correo.', 'warning')

    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/eliminar_visita_grupal/<int:id>')
@login_required
def eliminar_visita_grupal(id):
    visita = VisitaGrupal.query.get(id)
    if not visita:
        flash('❌ Visita no encontrada.', 'danger')
        return redirect(url_for('admin.dashboard'))

    if visita.estado not in ['cancelada', 'rechazada']:
        flash('❌ Solo puedes eliminar visitas canceladas o rechazadas.', 'danger')
        return redirect(url_for('admin.dashboard'))

    for estudiante in visita.estudiantes:
        db.session.delete(estudiante)
    db.session.delete(visita)
    db.session.commit()
    flash('✅ Visita eliminada correctamente.', 'success')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/generar_password', methods=['POST'])
@login_required
def generar_password():
    nueva = generar_password_segura(14)
    ahora_utc = datetime.utcnow()
    secret = AdminSecret.query.get(1)
    if secret:
        secret.password = nueva
        secret.created_at = ahora_utc
    else:
        secret = AdminSecret(id=1, password=nueva, created_at=ahora_utc)
        db.session.add(secret)
    db.session.commit()
    flash('🔐 Nueva contraseña generada.', 'success')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/descargar_historial')
@login_required
def descargar_historial():
    import io as _io
    zona  = pytz.timezone('America/Chihuahua')
    ahora = datetime.now(zona)
    rango = request.args.get('rango', default='30')
    tipo  = request.args.get('tipo',  default='todas')

    if rango == '7':
        inicio_rango = ahora - timedelta(days=7)
    elif rango == '30':
        inicio_rango = ahora - timedelta(days=30)
    elif rango == 'mes':
        inicio_rango = ahora.replace(day=1)
    elif rango == 'todo':
        inicio_rango = datetime.min.replace(tzinfo=zona)
    else:
        inicio_rango = ahora - timedelta(days=30)

    fecha_str    = ahora.strftime("%Y-%m-%d_%H-%M")
    buffer       = _io.BytesIO()
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def _datos_individuales():
        rows = []
        for c in Cita.query.all():
            try:
                fecha = datetime.strptime(c.fecha_hora, "%d/%m/%Y %I:%M %p")
            except ValueError:
                fecha = datetime.strptime(c.fecha_hora, "%Y-%m-%d %H:%M")
            fecha = zona.localize(fecha)
            if fecha < ahora and fecha >= inicio_rango and c.estado != 'cancelada':
                rows.append({
                    'ID':              c.id,
                    'Nombre':          c.nombre,
                    'Correo':          c.correo,
                    'Teléfono':        c.telefono,
                    'Edad':            c.edad,
                    'Sexo':            c.sexo,
                    'Ciudad':          c.ciudad,
                    'Estado':          c.estado_republica,
                    'Institución':     c.institucion,
                    'Nivel Académico': c.nivel_educativo,
                    'Fecha y Hora':    c.fecha_hora,
                })
        return rows

    def _datos_grupales():
        rows = []
        for v in VisitaGrupal.query.filter(
            VisitaGrupal.fecha_confirmada.isnot(None),
            VisitaGrupal.estado == 'aceptada'
        ).all():
            try:
                fecha_v = datetime.strptime(v.fecha_confirmada, "%d/%m/%Y %I:%M %p")
            except ValueError:
                try:
                    fecha_v = datetime.strptime(v.fecha_confirmada, "%d/%m/%Y %H:%M")
                except ValueError:
                    continue
            fecha_v = zona.localize(fecha_v)
            if not (fecha_v < ahora and fecha_v >= inicio_rango):
                continue

            if v.estudiantes:
                for est in v.estudiantes:
                    rows.append({
                        'Visita ID':       v.id,
                        'Institución':     v.institucion,
                        'Nivel Académico': v.nivel,
                        'Ciudad':          v.ciudad or '—',
                        'Estado':          v.estado_republica or '—',
                        'Encargado':       v.encargado,
                        'Fecha de visita': v.fecha_confirmada,
                        'Nombre alumno':   est.nombre,
                        'Edad alumno':     est.edad,
                        'Sexo alumno':     est.sexo,
                    })
            else:
                rows.append({
                    'Visita ID':       v.id,
                    'Institución':     v.institucion,
                    'Nivel Académico': v.nivel,
                    'Ciudad':          v.ciudad or '—',
                    'Estado':          v.estado_republica or '—',
                    'Encargado':       v.encargado,
                    'Fecha de visita': v.fecha_confirmada,
                    'Nombre alumno':   '(sin lista subida)',
                    'Edad alumno':     None,
                    'Sexo alumno':     None,
                })
        return rows

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        if tipo in ('todas', 'individual'):
            df_ind = pd.DataFrame(_datos_individuales())
            if df_ind.empty:
                df_ind = pd.DataFrame(columns=[
                    'ID','Nombre','Correo','Teléfono','Edad','Sexo',
                    'Ciudad','Estado','Institución','Nivel Académico','Fecha y Hora'
                ])
            df_ind.to_excel(writer, sheet_name='Citas Individuales', index=False)

        if tipo in ('todas', 'grupal'):
            df_grp = pd.DataFrame(_datos_grupales())
            if df_grp.empty:
                df_grp = pd.DataFrame(columns=[
                    'Visita ID','Institución','Nivel Académico','Ciudad','Estado',
                    'Encargado','Fecha de visita',
                    'Nombre alumno','Edad alumno','Sexo alumno'
                ])
            df_grp.to_excel(writer, sheet_name='Visitas Grupales', index=False)

    buffer.seek(0)

    if tipo == 'individual':
        nombre_archivo = f"historial_individual_{fecha_str}.xlsx"
    elif tipo == 'grupal':
        nombre_archivo = f"historial_grupal_{fecha_str}.xlsx"
    else:
        nombre_archivo = f"historial_completo_{fecha_str}.xlsx"

    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
    response.headers['Content-Type'] = content_type
    return response


@admin_bp.route('/subir_excel_visita/<int:id>', methods=['POST'])
@login_required
def subir_excel_visita(id):
    from app.models import EstudianteGrupal
    import openpyxl as _openpyxl

    visita = VisitaGrupal.query.get(id)
    if not visita:
        flash('❌ Visita no encontrada.', 'danger')
        return redirect(url_for('admin.dashboard'))

    archivo = request.files.get('excel_estudiantes')
    if not archivo or not (archivo.filename.endswith('.xlsx') or archivo.filename.endswith('.xls')):
        flash('❌ Debes subir un archivo Excel válido (.xlsx o .xls).', 'danger')
        return redirect(url_for('admin.dashboard'))

    try:
        if archivo.filename.endswith('.xls'):
            import io as _io
            df_raw = pd.read_excel(_io.BytesIO(archivo.read()), header=None, engine='xlrd')
            wb = _openpyxl.Workbook()
            ws = wb.active
            for row in df_raw.itertuples(index=False):
                ws.append(list(row))
        else:
            wb = _openpyxl.load_workbook(archivo)
            ws = wb.active

        fila_encabezado = None
        for row in ws.iter_rows():
            vals = [str(c.value).strip() if c.value else '' for c in row]
            if vals[0] == 'No.' and 'Nombre completo' in vals[1]:
                fila_encabezado = row[0].row
                break

        if fila_encabezado is None:
            flash('❌ No se encontró la tabla de alumnos. '
                  'No modifiques los encabezados del Excel.', 'danger')
            return redirect(url_for('admin.dashboard'))

        for est in visita.estudiantes:
            db.session.delete(est)
        db.session.flush()

        zona    = pytz.timezone('America/Chihuahua')
        ahora   = datetime.now(zona).strftime("%d/%m/%Y %I:%M %p")
        guardados = 0

        for row in ws.iter_rows(min_row=fila_encabezado + 1, values_only=True):
            _, nombre, edad, sexo = (row[i] if i < len(row) else None for i in range(4))
            nombre = str(nombre).strip() if nombre else ''
            if not nombre or nombre.lower() in ('none', '—', '-', ''):
                continue
            try:
                edad_int = int(edad) if edad else None
            except (ValueError, TypeError):
                edad_int = None
            from app.models import EstudianteGrupal as _EG
            db.session.add(_EG(
                nombre=nombre, edad=edad_int,
                sexo=str(sexo).strip() if sexo else None,
                hora_registro=ahora, visita_id=visita.id,
            ))
            guardados += 1

        db.session.commit()
        flash(f'✅ Excel procesado: {guardados} alumno(s) registrado(s) '
              f'para la visita de {visita.institucion}.', 'success')

    except Exception as e:
        db.session.rollback()
        print(f"[EXCEL] Error: {e}")
        flash('❌ Error al procesar el Excel. Verifica que no esté dañado.', 'danger')

    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/db-ping')
def db_ping():
    from sqlalchemy import text
    try:
        db.session.execute(text("select 1"))
        return "ok", 200
    except Exception as e:
        return f"db error: {e}", 500
